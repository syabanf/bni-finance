#!/usr/bin/env python3
"""Membangun workbook eksekusi QA dari docs/QA-E2E.md.

    python3 scripts/build-qa-xlsx.py

Sumbernya markdown, bukan spreadsheet, supaya perubahan skenario bisa di-review
sebagai diff di pull request. Excel-nya artefak turunan — ia masuk test-report/
yang gitignored, dan dibangun ulang kapan pun dibutuhkan.

Butuh openpyxl:  pip3 install openpyxl
"""
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit('butuh openpyxl — jalankan: pip3 install openpyxl')

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / 'docs' / 'QA-E2E.md'
OUT = ROOT / 'test-report' / 'skenario-qa-e2e.xlsx'

COLS = ['ID', 'Modul', 'Pri', 'Judul', 'Prasyarat', 'Langkah', 'Hasil yang diharapkan', 'Jenis']
NAVY, WHITE, GREY, ZEBRA = '00205B', 'FFFFFF', 'F2F4F7', 'FAFBFC'
PRI_BG = {'P1': 'FBE3E3', 'P2': 'FFF4E0', 'P3': 'EAF1FB'}
_thin = Side(style='thin', color='D6DBE1')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Baris skenario dikenali dari pola ID di kolom pertama — bukan dari posisi,
# supaya menambah paragraf di atas tabel tidak merusak parser.
ROW = re.compile(r'^\|\s*([A-Z]+-\d+)\s*\|')


def parse(md: str):
    rows = []
    for line in md.splitlines():
        if not ROW.match(line):
            continue
        cells = [c.strip().replace('<br>', '\n').replace('\\|', '|')
                 for c in line.strip().strip('|').split('|')]
        if len(cells) != len(COLS):
            sys.exit(f'baris {cells[0]!r} punya {len(cells)} kolom, harus {len(COLS)} — '
                     f'periksa tabel di {SRC.name}')
        rows.append(cells)
    if not rows:
        sys.exit(f'tidak ada baris skenario di {SRC}')
    ids = [r[0] for r in rows]
    dupes = {i for i in ids if ids.count(i) > 1}
    if dupes:
        sys.exit(f'ID ganda: {", ".join(sorted(dupes))}')
    return rows


def build(rows):
    wb = Workbook()

    # --- Panduan ---
    ws = wb.active
    ws.title = 'Panduan'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 96
    t = ws.cell(row=1, column=2, value='Skenario QA End-to-End — BNI Finance Hub')
    t.font = Font(bold=True, size=15, color=NAVY)
    ws.cell(row=2, column=2,
            value=f'{len(rows)} skenario. Sumber: docs/QA-E2E.md — ubah di sana, '
                  f'lalu bangun ulang dengan scripts/build-qa-xlsx.py').font = \
        Font(size=9, italic=True, color='5A6672')

    guide = [
        ('Cara pakai', 'Sheet "Skenario" adalah daftar eksekusi. Isi kolom Status, Tester, '
                       'Tanggal, dan Temuan saat menjalankan. Gunakan autofilter pada kolom '
                       'Modul atau Pri untuk membagi pekerjaan.'),
        ('Prioritas', 'P1 = menghalangi rilis; harus hijau sebelum ship. P2 = penting, boleh '
                      'menyusul dengan catatan. P3 = pelengkap.'),
        ('Jenis', 'Happy path · Negatif · Keamanan · Balapan (concurrency) · Integrasi · Beban.'),
        ('Prasyarat', 'backend/.env terisi (npm run setup) · Postgres termigrasi · db/seed.sql '
                      'dimuat · kredensial Paper.id staging untuk skenario Integrasi · dua akun: '
                      'satu admin, satu user biasa.'),
        ('Kanal pengiriman', 'Skenario Integrasi yang mengirim email/WhatsApp menghubungi kontak '
                             'SUNGGUHAN di data. Arahkan seluruh kontak ke tim penguji lebih '
                             'dulu, atau matikan kanalnya.'),
        ('Nomor invoice', 'Paper.id membakar nomor invoice permanen begitu dipakai. Membersihkan '
                          'database lokal tidak mengembalikannya — lihat X-02.'),
        ('Catatan regresi', 'Beberapa skenario memuat CATATAN berisi bug yang pernah benar-benar '
                            'terjadi. Skenario itu penjaga regresi dan paling layak dijalankan '
                            'lebih dulu.'),
    ]
    r = 4
    for head, body in guide:
        h = ws.cell(row=r, column=2, value=head)
        h.font = Font(bold=True, size=10, color=NAVY)
        h.alignment = Alignment(vertical='top', wrap_text=True)
        c = ws.cell(row=r, column=3, value=body)
        c.font = Font(size=10, color='36414D')
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 42
        for cc in (2, 3):
            ws.cell(row=r, column=cc).border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=2, value='Rekap').font = Font(bold=True, size=12, color=NAVY)
    r += 1
    for i, lab in enumerate(('Modul', 'Total (P1 / P2 / P3)')):
        c = ws.cell(row=r, column=2 + i, value=lab)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.border = BORDER
    r += 1
    mods = list(dict.fromkeys(x[1] for x in rows))
    for m in mods:
        sub = [x for x in rows if x[1] == m]
        n = {p: sum(1 for x in sub if x[2] == p) for p in ('P1', 'P2', 'P3')}
        ws.cell(row=r, column=2, value=m).font = Font(size=10, bold=True)
        ws.cell(row=r, column=3,
                value=f"{len(sub)}   ({n['P1']} / {n['P2']} / {n['P3']})").font = Font(size=10)
        for cc in (2, 3):
            ws.cell(row=r, column=cc).border = BORDER
        r += 1
    tot = {p: sum(1 for x in rows if x[2] == p) for p in ('P1', 'P2', 'P3')}
    ws.cell(row=r, column=2, value='TOTAL').font = Font(bold=True, size=10, color=NAVY)
    ws.cell(row=r, column=3,
            value=f"{len(rows)}   ({tot['P1']} / {tot['P2']} / {tot['P3']})").font = \
        Font(bold=True, size=10, color=NAVY)
    for cc in (2, 3):
        ws.cell(row=r, column=cc).border = BORDER
        ws.cell(row=r, column=cc).fill = PatternFill('solid', fgColor=GREY)

    # --- Skenario ---
    ws2 = wb.create_sheet('Skenario')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 2
    ws2.cell(row=1, column=2, value='Daftar eksekusi').font = Font(bold=True, size=15, color=NAVY)
    ws2.cell(row=2, column=2,
             value='Isi Status / Tester / Tanggal / Temuan saat menjalankan.').font = \
        Font(size=9, italic=True, color='5A6672')

    headers = COLS + ['Status', 'Tester', 'Tanggal', 'Temuan']
    widths = [10, 17, 6, 34, 30, 44, 52, 13, 13, 12, 12, 34]
    for i, (lab, w) in enumerate(zip(headers, widths), start=2):
        c = ws2.cell(row=4, column=i, value=lab)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER
        ws2.column_dimensions[c.column_letter].width = w
    ws2.row_dimensions[4].height = 26
    ws2.freeze_panes = 'E5'

    for i, row in enumerate(rows):
        rr = 5 + i
        for j, v in enumerate(row + ['', '', '', '']):
            c = ws2.cell(row=rr, column=2 + j, value=v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.font = Font(size=9)
            if i % 2 and c.column != 4:
                c.fill = PatternFill('solid', fgColor=ZEBRA)
        ws2.cell(row=rr, column=2).font = Font(size=9, bold=True, color=NAVY)
        ws2.cell(row=rr, column=5).font = Font(size=9, bold=True)
        pc = ws2.cell(row=rr, column=4)
        pc.fill = PatternFill('solid', fgColor=PRI_BG.get(row[2], ZEBRA))
        pc.font = Font(size=9, bold=True)
        pc.alignment = Alignment(horizontal='center', vertical='center')
        longest = max(row[5].count('\n'), row[6].count('\n'))
        ws2.row_dimensions[rr].height = max(46, 12 * longest + 34)

    dv = DataValidation(type='list', formula1='"Belum,Lulus,Gagal,Blocked,N/A"', allow_blank=True)
    ws2.add_data_validation(dv)
    dv.add(f'J5:J{4 + len(rows)}')
    ws2.auto_filter.ref = f'B4:M{4 + len(rows)}'
    return wb


def main():
    if not SRC.exists():
        sys.exit(f'tidak ada {SRC}')
    rows = parse(SRC.read_text(encoding='utf-8'))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    build(rows).save(OUT)
    print(f'{OUT.relative_to(ROOT)} — {len(rows)} skenario')


if __name__ == '__main__':
    main()
