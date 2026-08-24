#!/usr/bin/env python3
"""Membuat ulang berkas uji XLSX di direktori ini.

    python3 backend/internal/importer/testdata/generate.py

DUA BENTUK, dan keduanya perlu — inilah alasan berkas ini ada.

XLSX menyimpan teks dengan dua cara yang berbeda:

    inlineStr        teksnya langsung di dalam sel
    sharedStrings    sel hanya memuat INDEKS ke tabel teks terpisah

openpyxl menulis bentuk pertama. **Microsoft Excel dan Google Sheets menulis
bentuk kedua.** Jadi berkas uji yang seluruhnya dibuat openpyxl tidak pernah
menyentuh jalur yang justru akan dilewati hampir semua berkas nyata.

Itu bukan kekhawatiran teoretis: ketahuan saat menyabotase pembaca tabel teks
bersama dan seluruh tes tetap hijau. Tanpa fixture bentuk kedua, kerusakan pada
jalur itu tidak akan pernah terlihat sampai ada orang mengunggah berkas dari
Excel dan mendapati seluruh kolom teksnya berisi "0", "1", "2".

Berkas sharedStrings di bawah karena itu disusun langsung dengan zipfile,
mengikuti bentuk yang benar-benar dihasilkan Excel.
"""

import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

OUT = Path(__file__).parent

# --- bentuk inlineStr, lewat openpyxl ---------------------------------------
from openpyxl import Workbook  # noqa: E402

wb = Workbook()
ws = wb.active
ws.append(["id", "chapter_id", "name", "email", "phone", "company"])
ws.append(["mem-101", "ch-garuda", "Andi Pratama", "andi@contoh.id", "081234567890", "PT Contoh"])
ws.append(["mem-102", "ch-nusantara", "Budi Rahayu", "budi@contoh.id", "081234567891", "CV Contoh"])
wb.save(OUT / "member-normal.xlsx")

# Sel kosong di tengah. Excel TIDAK menulis sel kosong sama sekali, jadi berkas
# ini menguji apakah pembacanya menggeser kolom setelahnya.
wb = Workbook()
ws = wb.active
ws.append(["id", "chapter_id", "name", "email"])
ws["A2"] = "mem-201"; ws["C2"] = "Tanpa Chapter"; ws["D2"] = "a@contoh.id"
ws["A3"] = "mem-202"; ws["B3"] = "ch-garuda"; ws["C3"] = "Tanpa Email"
wb.save(OUT / "member-sel-kosong.xlsx")

# Teks berformat campuran dalam satu sel.
wb = Workbook()
ws = wb.active
ws.append(["id", "name"])
ws.append(["mem-301", "Nama Panjang Sekali Yang Sebagian Ditebalkan"])
wb.save(OUT / "member-richtext.xlsx")


# --- bentuk sharedStrings, disusun langsung ---------------------------------
def xlsx_shared_strings(path: Path, rows: list[list[str]]) -> None:
    """Tulis XLSX yang memakai sharedStrings, seperti yang dihasilkan Excel."""
    # Tabel teks: tiap teks unik satu kali, sel menyimpan indeksnya.
    unik: list[str] = []
    indeks: dict[str, int] = {}
    for baris in rows:
        for sel in baris:
            if sel != "" and sel not in indeks:
                indeks[sel] = len(unik)
                unik.append(sel)

    sst = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{sum(1 for b in rows for s in b if s != "")}" uniqueCount="{len(unik)}">'
        + "".join(f"<si><t>{escape(s)}</t></si>" for s in unik)
        + "</sst>"
    )

    def huruf(i: int) -> str:
        s = ""
        i += 1
        while i:
            i, sisa = divmod(i - 1, 26)
            s = chr(65 + sisa) + s
        return s

    sel_xml = []
    for r, baris in enumerate(rows, start=1):
        cells = "".join(
            f'<c r="{huruf(c)}{r}" t="s"><v>{indeks[s]}</v></c>'
            for c, s in enumerate(baris)
            if s != ""
        )
        sel_xml.append(f'<row r="{r}">{cells}</row>')

    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData>" + "".join(sel_xml) + "</sheetData></worksheet>"
    )

    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    workbook = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    wb_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        "</Relationships>"
    )

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("xl/workbook.xml", workbook)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/sharedStrings.xml", sst)
        z.writestr("xl/worksheets/sheet1.xml", sheet)


xlsx_shared_strings(
    OUT / "member-sharedstrings.xlsx",
    [
        ["id", "chapter_id", "name", "email"],
        ["mem-401", "ch-garuda", "Citra Melati", "citra@contoh.id"],
        # Baris dengan kolom kosong di tengah, pada bentuk sharedStrings.
        ["mem-402", "", "Dian Kusuma", "dian@contoh.id"],
        # Teks yang berulang: dua baris memakai chapter yang sama, sehingga
        # tabel teksnya benar-benar dipakai ulang seperti pada berkas nyata.
        ["mem-403", "ch-garuda", "Eko Saputra", "eko@contoh.id"],
    ],
)

print("berkas uji dibuat ulang di", OUT)
